"""Generate data/patient_data_template.xlsx — the manual data entry template."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Color scheme ──────────────────────────────────────────────────────────────
# Navy   = join key (pt_uuid, report_uuid) — links rows across sheets
# Orange = canonical normalized field (gene, variant_type, result_summary)
# Green  = raw source field — verbatim text, never edited after entry
# Gray   = plain structured field
FILLS = {
    "id":        PatternFill("solid", fgColor="1F3864"),
    "canonical": PatternFill("solid", fgColor="843C0C"),
    "raw":       PatternFill("solid", fgColor="375623"),
    "plain":     PatternFill("solid", fgColor="404040"),
}
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
EXAMPLE_FONT = Font(italic=True, color="AAAAAA", size=9)


def add_sheet(wb, name, columns, example=None):
    """columns: list of (col_name, color_key). example: list of values for row 2."""
    ws = wb.create_sheet(name)
    for i, (col_name, color_key) in enumerate(columns, 1):
        c = ws.cell(row=1, column=i, value=col_name)
        c.fill = FILLS[color_key]
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(i)].width = max(len(col_name) + 4, 20)
    if example:
        for i, val in enumerate(example, 1):
            c = ws.cell(row=2, column=i, value=val)
            c.font = EXAMPLE_FONT
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    return ws


# ── _legend ───────────────────────────────────────────────────────────────────
leg = wb.create_sheet("_legend")
leg.column_dimensions["A"].width = 35
leg.column_dimensions["B"].width = 65
legend_rows = [
    ("Column type",                          "Meaning",                                                                   None),
    ("pt_uuid / report_uuid",                "Join keys — link rows across sheets into the same MongoDB document",       "id"),
    ("gene, variant_type, result_summary",   "Canonical normalized fields — must be consistent across ALL sources",      "canonical"),
    ("raw_*",                                "Verbatim text from the source — never edit after entry",                   "raw"),
    ("All other columns",                    "Plain structured fields",                                                   "plain"),
]
for r, (col_type, meaning, key) in enumerate(legend_rows, 1):
    c1 = leg.cell(row=r, column=1, value=col_type)
    leg.cell(row=r, column=2, value=meaning)
    if key:
        c1.fill = FILLS[key]
        c1.font = HEADER_FONT
    else:
        c1.font = Font(bold=True)
leg.row_dimensions[1].height = 20


# ── _conventions ──────────────────────────────────────────────────────────────
conv = wb.create_sheet("_conventions")
conv.column_dimensions["A"].width = 25
conv.column_dimensions["B"].width = 70

conv_header = conv.cell(row=1, column=1, value="Field")
conv_header.fill = FILLS["plain"]
conv_header.font = HEADER_FONT
conv.cell(row=1, column=2, value="Allowed values / notes").fill = FILLS["plain"]
conv.cell(row=1, column=2).font = HEADER_FONT

conventions = [
    ("variant_type",    "somatic_mutation | germline | structural_variant | fusion | expression | "
                        "tumor_biomarker | pertinent_negative | immunotherapy_marker | hla | indeterminate"),
    ("source",          "tempus | caris | ambry | amc_ngs | ogm | pml_rara  (add new values as needed)"),
    ("gene",            "Use HGNC approved symbol: EGFR, ERBB2, TP53, KRAS, etc. "
                        "For fusions: PML/RARA. For biomarkers: TMB, MSI, PD-L1, LOH, HRD, MMR."),
    ("result_summary",  "Short human-readable string. Examples: p.L858R, exon 19 del, negative, "
                        "3.2 mut/Mb, MSI-High, positive 3+ 95%"),
    ("sex",             "male | female | other"),
    ("pt_uuid",         "Auto-incrementing integer. Assign sequentially: 0, 1, 2 …"),
    ("report_uuid",     "Auto-incrementing integer across ALL report_metadata rows: 0, 1, 2 …"),
]
for r, (field, notes) in enumerate(conventions, 2):
    conv.cell(row=r, column=1, value=field).font = Font(bold=True)
    conv.cell(row=r, column=2, value=notes)
conv.freeze_panes = "A2"
conv.row_dimensions[1].height = 28


# ── pt_general ────────────────────────────────────────────────────────────────
add_sheet(wb, "pt_general", [
    ("pt_uuid",           "id"),
    ("mrn",               "plain"),
    ("first_name",        "plain"),
    ("last_name",         "plain"),
    ("dob",               "plain"),
    ("sex",               "plain"),
    ("entity",            "plain"),
    ("primary_dx",        "plain"),
    ("metastasis_sites",  "plain"),
], example=[
    0, 302939, "Larry", "Corum", None, None, "AMC", "mid-rectal adenocarcinoma", None,
])


# ── report_metadata ───────────────────────────────────────────────────────────
add_sheet(wb, "report_metadata", [
    ("report_uuid",    "id"),
    ("pt_uuid",        "id"),
    ("source",         "plain"),
    ("test_name",      "plain"),
    ("accession_no",   "plain"),
    ("physician",      "plain"),
    ("specimen_type",  "plain"),
    ("date_collected", "plain"),
    ("date_received",  "plain"),
    ("date_completed", "plain"),
    ("obtained_from",  "plain"),
    ("link",           "plain"),
    ("notes",          "plain"),
], example=[
    0, 1, "tempus", "xT CDx", "TL-26-LOVTO77643", "Erin Cobain",
    "tumor/blood", "2026-01-08", "2026-02-25", "2026-03-07",
    "Erin Cobain Email", "https://...", None,
])


# ── tempus_findings ───────────────────────────────────────────────────────────
add_sheet(wb, "tempus_findings", [
    ("pt_uuid",                  "id"),
    ("report_uuid",              "id"),
    ("gene",                     "canonical"),
    ("variant_type",             "canonical"),
    ("result_summary",           "canonical"),
    ("raw_biomarker",            "raw"),
    ("raw_result",               "raw"),
    ("raw_category",             "raw"),
    ("raw_nucleotide_type",      "raw"),
    ("raw_therapies_current_dx", "raw"),
    ("raw_therapies_other",      "raw"),
    ("raw_trials",               "raw"),
], example=[
    1, 0, "ERBB2", "somatic_mutation", "p.T733I 53.2% VAF",
    "ERBB2 (HER2) p.T733I", "53.2% VAF", "somatic - potentially actionable",
    "DNA", None, "Neratinib, Neratinib + Fulvestrant", "2 trials",
])


# ── caris_findings ────────────────────────────────────────────────────────────
# Flattens all Caris report sections into one sheet.
# Specimen-level fields repeat on every finding row for that report (parsed by report_uuid).
# raw_section tracks which part of the Caris report each row came from.
add_sheet(wb, "caris_findings", [
    # ── join keys ──
    ("pt_uuid",                    "id"),
    ("report_uuid",                "id"),
    # ── canonical normalized ──
    ("gene",                       "canonical"),
    ("variant_type",               "canonical"),
    ("result_summary",             "canonical"),
    # ── specimen info (Caris-specific, repeats per row) ──
    ("raw_specimen_id",            "raw"),
    ("raw_primary_tumor_site",     "raw"),
    ("raw_specimen_site",          "raw"),
    ("raw_specimen_collected",     "raw"),
    ("raw_test_report_date",       "raw"),
    ("raw_completion_of_addendum", "raw"),
    ("raw_ordered_by_location",    "raw"),
    # ── finding fields ──
    ("raw_section",                "raw"),   # therapy_associations | cancer_type_biomarkers |
                                             # genomic_signatures | pathogenic_alterations |
                                             # hla | ihc | indeterminate | intermediate_can
    ("raw_biomarker",              "raw"),
    ("raw_method",                 "raw"),
    ("raw_analyte",                "raw"),
    ("raw_result",                 "raw"),
    ("raw_benefit",                "raw"),
    ("raw_therapy_assoc",          "raw"),
    ("raw_biomarker_level",        "raw"),
    ("raw_protein_alteration",     "raw"),
    ("raw_exon",                   "raw"),
    ("raw_dna_alteration",         "raw"),
    ("raw_frequency_pct",          "raw"),
    ("raw_genotype",               "raw"),   # HLA rows
    ("raw_hla_class",              "raw"),   # HLA rows
], example=[
    2, 1, "ROS1", "fusion", "CD74-ROS1 pathogenic fusion",
    "SP26-02346A1", "lower lobe, lung", "lower lobe, lung",
    "2026-02-17", "2026-03-17", "2026-03-21",
    "University of Michigan Health West - Cancer Center at the Village",
    "pathogenic_alterations", "ROS1", "Seq", "RNA-Tumor",
    "Pathogenic Fusion | CD74=ROS1", None, None, None,
    "CD74-ROS1", 34, "NA", "NA", None, None,
])


# ── ambry_findings ────────────────────────────────────────────────────────────
add_sheet(wb, "ambry_findings", [
    ("pt_uuid",                        "id"),
    ("report_uuid",                    "id"),
    ("gene",                           "canonical"),
    ("variant_type",                   "canonical"),
    ("result_summary",                 "canonical"),
    ("raw_pathogenic_mutations",       "raw"),
    ("raw_vus",                        "raw"),
    ("raw_gross_deletions_dups",       "raw"),
    ("raw_summary",                    "raw"),
], example=[
    2, 2, None, "germline", "negative — no significant variants",
    "None Detected", "None Detected", "None Detected",
    "Negative: No clinically significant variants detected",
])


# ── amc_ngs_findings ──────────────────────────────────────────────────────────
add_sheet(wb, "amc_ngs_findings", [
    # ── join keys ──
    ("pt_uuid",                      "id"),
    ("report_uuid",                  "id"),
    # ── canonical normalized ──
    ("gene",                         "canonical"),
    ("variant_type",                 "canonical"),
    ("result_summary",               "canonical"),
    # ── specimen info (AMC-specific, repeats per row) ──
    ("raw_specimen_id",              "raw"),
    ("raw_block_id",                 "raw"),
    ("raw_body_site",                "raw"),
    # ── finding fields ──
    ("raw_finding_level",            "raw"),
    ("raw_variant_name",             "raw"),
    ("raw_dna_change",               "raw"),
    ("raw_amino_acid_change",        "raw"),
    ("raw_transcript",               "raw"),
    ("raw_interpretation",           "raw"),
    ("raw_therapeutic_implications", "raw"),
    ("raw_pertinent_negatives",      "raw"),
], example=[
    3, 3, "EGFR", "somatic_mutation", "exon 19 del E746_A750del",
    "F91209848", "SU-26-14583-A1", "Left lower lobe",
    "Level 1 FDA-recognized biomarker", "EGFR exon 19 deletion",
    "c.2236_2250del", "E746_A750del", "NM_005228.3 hg19 chr7:55242465",
    "EGFR E746_A750del: constitutive oncogenic activation of EGFR...",
    "Osimertinib indicated for first-line treatment of NSCLC...",
    "Absence of BRAF mutation, KRAS mutation, ALK rearrangement...",
])


# ── ogm_findings ──────────────────────────────────────────────────────────────
add_sheet(wb, "ogm_findings", [
    ("pt_uuid",                "id"),
    ("report_uuid",            "id"),
    ("gene",                   "canonical"),
    ("variant_type",           "canonical"),
    ("result_summary",         "canonical"),
    ("raw_selected_results",   "raw"),
    ("raw_interpretation",     "raw"),
    ("raw_iscn_karyotype",     "raw"),
    ("raw_additional_results", "raw"),
], example=[
    4, 4, None, "structural_variant", "no significant alterations detected",
    "No clinically significant copy number alterations or structural variants were detected.",
    "Optical genome mapping analysis did not identify any clinically significant genomic alterations.",
    "ogm (X,1-22)x2", None,
])


# ── pml_rara_findings ─────────────────────────────────────────────────────────
add_sheet(wb, "pml_rara_findings", [
    ("pt_uuid",           "id"),
    ("report_uuid",       "id"),
    ("gene",              "canonical"),
    ("variant_type",      "canonical"),
    ("result_summary",    "canonical"),
    ("raw_test_result",   "raw"),
    ("raw_interpretation","raw"),
], example=[
    4, 5, "PML/RARA", "fusion", "negative",
    "NEGATIVE for PML/RARA",
    "PML/RARA transcripts were not detected in this specimen.",
])


# ── tumor_biomarkers ──────────────────────────────────────────────────────────
add_sheet(wb, "tumor_biomarkers", [
    ("pt_uuid",            "id"),
    ("report_uuid",        "id"),
    ("gene",               "canonical"),   # use biomarker name: TMB, MSI, PD-L1 …
    ("variant_type",       "canonical"),   # always: tumor_biomarker
    ("result_summary",     "canonical"),
    ("raw_tmb",            "raw"),
    ("raw_msi",            "raw"),
    ("raw_pd_l1",          "raw"),
    ("raw_loh",            "raw"),
    ("raw_hrd",            "raw"),
    ("raw_mmr",            "raw"),
    ("raw_tumor_fraction", "raw"),
    ("raw_tumor_normal",   "raw"),
    ("raw_rna_expression", "raw"),
    ("raw_rna_fusion",     "raw"),
], example=[
    3, 3, "TMB", "tumor_biomarker", "1.6 mut/Mb",
    "1.6m/MB", "stable", "NA", "NA",
    "Not detected (<0.1, threshold 50)", "normal", "3.1", "negative",
    "overexpression of ERBB3, WT1, underexpression of CDKN2A/B", "negative",
])


out = "data/patient_data_template.xlsx"
wb.save(out)
print(f"Saved → {out}")
print(f"Sheets: {wb.sheetnames}")
