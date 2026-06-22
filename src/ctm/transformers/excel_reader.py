"""Read patient_data_template.xlsx → normalized Patient, ReportMetadata, Finding instances."""
from pathlib import Path

import openpyxl

from ..schemas.raw.models import RawPatientGeneral, RawReportMetadata
from ..schemas.raw.normalized import Finding, Patient, ReportMetadata
from .normalize_manual import (
    SHEET_NORMALIZERS,
    normalize_patient,
    normalize_report_metadata,
)


def _sheet_rows(ws) -> list[dict]:
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def read_and_normalize(
    path: Path,
    pt_uuid_filter: int | None = None,
) -> tuple[list[Patient], list[ReportMetadata], list[Finding]]:
    """Read Excel workbook → (patients, report_metadata, findings), all normalized.

    pt_uuid_filter: if set, only rows for that patient are returned.
    Rows that fail validation are skipped with a printed warning.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Patients ──────────────────────────────────────────────────────────────
    patients: list[Patient] = []
    if "pt_general" in wb.sheetnames:
        for row in _sheet_rows(wb["pt_general"]):
            if row.get("pt_uuid") is None:
                continue
            if pt_uuid_filter is not None and row["pt_uuid"] != pt_uuid_filter:
                continue
            try:
                patients.append(normalize_patient(RawPatientGeneral.model_validate(row)))
            except Exception as exc:
                print(f"  Warning: pt_general row skipped — {exc}")

    valid_pt_uuids = {p.pt_uuid for p in patients}

    # ── Report metadata ────────────────────────────────────────────────────────
    metadata: list[ReportMetadata] = []
    if "report_metadata" in wb.sheetnames:
        for row in _sheet_rows(wb["report_metadata"]):
            if row.get("report_uuid") is None:
                continue
            if row.get("pt_uuid") not in valid_pt_uuids:
                continue
            try:
                metadata.append(
                    normalize_report_metadata(RawReportMetadata.model_validate(row))
                )
            except Exception as exc:
                print(f"  Warning: report_metadata row skipped — {exc}")

    report_source: dict[int, str] = {m.report_uuid: m.source for m in metadata}

    # ── Findings (all source sheets) ──────────────────────────────────────────
    findings: list[Finding] = []
    for sheet_name, (raw_cls, norm_fn) in SHEET_NORMALIZERS.items():
        if sheet_name not in wb.sheetnames:
            continue
        for row in _sheet_rows(wb[sheet_name]):
            if row.get("pt_uuid") is None:
                continue
            if row.get("pt_uuid") not in valid_pt_uuids:
                continue
            try:
                raw = raw_cls.model_validate(row)
                source = report_source.get(
                    raw.report_uuid,
                    sheet_name.replace("_findings", ""),
                )
                findings.append(norm_fn(raw, source=source))
            except Exception as exc:
                print(f"  Warning: {sheet_name} row skipped — {exc}")

    return patients, metadata, findings
